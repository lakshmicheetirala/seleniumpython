import openpyxl

path = r"E:\selenium\registration_data.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.get_sheet_by_name("registration_data")

rows = sheet.max_row
cols = sheet.max_column

print("rows===================>", rows)
print("cols================>", cols)

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end = " ")

    print()